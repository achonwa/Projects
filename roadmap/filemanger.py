import os
import mimetypes
import tkinter as tk
from tkinter import filedialog, messagebox
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
from pydub import AudioSegment
import cv2
import zipfile
import sqlite3
import joblib

def read_binary_file(file_path):
    try:
        with open(file_path, "rb") as binary_file:
            data = binary_file.read()
            return data
    except FileNotFoundError:
        messagebox.showerror("Error", f"File '{file_path}' not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Error reading file: {e}")

def get_file_type(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    return mime_type

def extract_text_from_pdf(file_path):
    reader = PdfReader(file_path)
    content = ""
    for page in reader.pages:
        content += page.extract_text() + "\n"
    return content

def extract_text_from_docx(file_path):
    doc = Document(file_path)
    content = ""
    for paragraph in doc.paragraphs:
        content += paragraph.text + "\n"
    return content

def read_excel_file(file_path):
    workbook = load_workbook(filename=file_path)
    content = ""
    for sheet_name in workbook.sheetnames:
        content += f"Sheet: {sheet_name}\n"
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            content += f"{row}\n"
    return content

def resize_image(file_path, output_path, size=(100, 100)):
    with Image.open(file_path) as img:
        img = img.resize(size)
        img.save(output_path)
    return f"Image resized and saved to {output_path}"

def convert_audio_format(input_path, output_path):
    audio = AudioSegment.from_file(input_path)
    audio.export(output_path, format="mp3")
    return f"Audio converted and saved to {output_path}"

def extract_frames_from_video(file_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    video = cv2.VideoCapture(file_path)
    count = 0
    while True:
        ret, frame = video.read()
        if not ret:
            break
        frame_path = os.path.join(output_dir, f"frame_{count}.jpg")
        cv2.imwrite(frame_path, frame)
        count += 1
    video.release()
    return f"Extracted {count} frames to {output_dir}"

def extract_zip(file_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    with zipfile.ZipFile(file_path, "r") as zip_ref:
        zip_ref.extractall(output_dir)
    return f"Files extracted to {output_dir}"

def read_sqlite_db(file_path):
    conn = sqlite3.connect(file_path)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    conn.close()
    return f"Tables: {tables}"

def load_ml_model(file_path):
    model = joblib.load(file_path)
    return "Machine learning model loaded successfully."

def analyze_file():
    file_path = filedialog.askopenfilename()
    if not file_path:
        return

    mime_type = get_file_type(file_path)
    output = ""

    try:
        if mime_type == "application/pdf":
            output = extract_text_from_pdf(file_path)
        elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            output = extract_text_from_docx(file_path)
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            output = read_excel_file(file_path)
        elif mime_type and "image" in mime_type:
            output_path = "resized_image.jpg"
            output = resize_image(file_path, output_path)
        elif mime_type and "audio" in mime_type:
            output_path = "output_audio.mp3"
            output = convert_audio_format(file_path, output_path)
        elif mime_type and "video" in mime_type:
            output_dir = "video_frames"
            output = extract_frames_from_video(file_path, output_dir)
        elif mime_type == "application/zip":
            output_dir = "extracted_files"
            output = extract_zip(file_path, output_dir)
        elif mime_type == "application/x-sqlite3":
            output = read_sqlite_db(file_path)
        else:
            output = "Unsupported file type or unknown format."

    except Exception as e:
        output = f"Error processing file: {e}"

    display_output(output)

def display_output(output):
    output_text.delete("1.0", tk.END)
    output_text.insert(tk.END, output)

# Set up the GUI
root = tk.Tk()
root.title("Binary File Analyzer")

frame = tk.Frame(root)
frame.pack(pady=20)

select_file_button = tk.Button(frame, text="Select and Analyze File", command=analyze_file)
select_file_button.pack()

output_text = tk.Text(root, height=20, width=80)
output_text.pack(pady=10)

root.mainloop()
